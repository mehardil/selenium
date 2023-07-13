# import openpyxl

# # Load the Excel file
# filename = "like12.xlsx"
# workbook = openpyxl.load_workbook(filename)

# # Select the active sheet
# sheet = workbook.active

# # Iterate over the cells in column 12 and modify the values
# column_number = 22
# for cell in sheet.iter_cols(min_col=column_number, max_col=column_number):
#     for cell_value in cell:
#         if cell_value.value:
#             modified_value = cell_value.value.lstrip()
#             cell_value.value = modified_value

# # Save the modified workbook
# modified_filename = "modified_example55515.xlsx"
# workbook.save(modified_filename)













import openpyxl

# Load the Excel file
filename = "alldata-facebook-pull-request-main.xlsx"
workbook = openpyxl.load_workbook(filename)

# Select the active sheet
sheet = workbook.active

# Iterate over the cells in column 12 and check for empty values
column_number = 8
for cell in sheet.iter_cols(min_col=column_number, max_col=column_number):
    for cell_value in cell:
        if cell_value.value is None or cell_value.value == "":
            cell_value.value = "NaN"

# Save the modified workbook
modified_filename = "alldata-facebook-pull-data.xlsx"
workbook.save(modified_filename)

# import openpyxl

# # Load the Excel file
# filename = "like12.xlsx"
# workbook = openpyxl.load_workbook(filename)

# # Select the active sheet
# sheet = workbook.active

# # Iterate over the cells in column 12 and modify the values
# column_number = 22
# for cell in sheet.iter_cols(min_col=column_number, max_col=column_number):
#     for cell_value in cell:
#         if cell_value.value and len(cell_value.value) > 16:
#             words = cell_value.value.split(" ", 4)
#             modified_value = words[-1] if len(words) > 4 else ""
#             cell_value.value = modified_value

# # Save the modified workbook
# modified_filename = "modified_example1111.xlsx"
# workbook.save(modified_filename)






# import openpyxl

# # Load the Excel file
# filename = "removelike1.xlsx"
# workbook = openpyxl.load_workbook(filename)

# # Select the active sheet
# sheet = workbook.active

# # Iterate over the cells in column 12 and modify the values
# column_number = 20
# for cell in sheet.iter_cols(min_col=column_number, max_col=column_number):
#     for cell_value in cell:
#         if cell_value.value and len(cell_value.value) > 26:
#             words = cell_value.value.split(" ")
#             modified_value = " ".join(words[4:])
#             cell_value.value = modified_value

# # Save the modified workbook
# modified_filename = "removelike2.xlsx"
# workbook.save(modified_filename)













# import openpyxl

# # Load the Excel file
# filename = "x.xlsx"
# workbook = openpyxl.load_workbook(filename)

# # Select the active sheet
# sheet = workbook.active

# # Read data from column 12 and replace double spaces with commas
# column_number = 20
# for cell in sheet.iter_cols(min_col=column_number, max_col=column_number):
#     for cell_value in cell:
#         if cell_value.value and "  " in cell_value.value:
#             modified_value = cell_value.value.replace("  ", " ,")
#             cell_value.value = modified_value

# # Save the modified workbook
# modified_filename = "x.xlsx"
# workbook.save(modified_filename)
